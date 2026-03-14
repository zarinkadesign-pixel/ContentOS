"""
NOESIS — Deterministic Hybrid Control Framework for Frozen Neural Operators (DHCF-FNO)
Copyright (c) 2026 AMAImedia.com
All rights reserved.
astana_v3/export/excel_exporter.py
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import List, Optional

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Listing, ListingStatus


# ── Colour palette ────────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="1A3A5C")
_PENDING_FILL  = PatternFill("solid", fgColor="FFF3CD")
_POSTED_FILL   = PatternFill("solid", fgColor="D4EDDA")
_RENTED_FILL   = PatternFill("solid", fgColor="D6D8DB")
_REJECTED_FILL = PatternFill("solid", fgColor="F8D7DA")
_ALT_ROW_FILL  = PatternFill("solid", fgColor="F8F9FA")

# Border style shared across all data cells
_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Font definitions
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
_BODY_FONT   = Font(size=10, name="Arial")
_TITLE_FONT  = Font(bold=True, size=14, name="Arial", color="1A3A5C")

# Cell number formats
_MONEY_FORMAT = '#,##0 "₸"'
_DATE_FORMAT  = "DD.MM.YYYY HH:MM"


# ── Column definitions: (header, width, field_key) ────────────────────────────

_COLUMNS = [
    ("ID",              8,   "external_id"),
    ("Status",         12,   "status_label"),
    ("District",       14,   "district"),
    ("Address",        28,   "address"),
    ("Rooms",          10,   "rooms_label"),
    ("Area m²",        11,   "area"),
    ("Floor",           8,   "floor_str"),
    ("Price ₸/month",  15,   "price"),
    ("Deposit",        12,   "deposit"),
    ("Contract",       14,   "contract_months"),
    ("Amenities",      22,   "amenities"),
    ("Description",    35,   "description"),
    ("Photos",          9,   "photo_count"),
    ("Phone",          16,   "landlord_phone"),
    ("TG owner",       16,   "landlord_username"),
    ("Created",        18,   "created_at"),
    ("Approved",       18,   "approved_at"),
    ("Post count",     18,   "posted_count"),
    ("Rented",         18,   "rented_at"),
]

# Column indices (1-based) that contain monetary values
_MONEY_COL_INDICES = {8}
# Column indices that contain datetime values
_DATE_COL_INDICES  = {16, 17, 19}
# Column indices where text wrapping should be enabled
_WRAP_COL_INDICES  = {4, 11, 12}

# Status label -> fill mapping
_STATUS_FILL: dict = {
    ListingStatus.PENDING:  _PENDING_FILL,
    ListingStatus.POSTED:   _POSTED_FILL,
    ListingStatus.RENTED:   _RENTED_FILL,
    ListingStatus.REJECTED: _REJECTED_FILL,
}

# Human-readable status labels
_STATUS_LABELS: dict = {
    ListingStatus.PENDING:  "Pending",
    ListingStatus.APPROVED: "Approved",
    ListingStatus.POSTED:   "Published",
    ListingStatus.RENTED:   "Rented",
    ListingStatus.REJECTED: "Rejected",
    ListingStatus.PAUSED:   "Paused",
}


# ── Row value extractor ───────────────────────────────────────────────────────

def _listing_to_row(lst: Listing) -> list:
    """Converts a Listing into an ordered list of cell values for the sheet."""
    if lst.rooms == 0:
        rooms_label = "Studio"
    elif lst.rooms:
        rooms_label = f"{lst.rooms}-room"
    else:
        rooms_label = "—"

    if lst.floor and lst.total_floors:
        floor_str = f"{lst.floor}/{lst.total_floors}"
    elif lst.floor:
        floor_str = str(lst.floor)
    else:
        floor_str = "—"

    return [
        lst.external_id,
        _STATUS_LABELS.get(lst.status, lst.status.value),
        lst.district,
        lst.address,
        rooms_label,
        lst.area,
        floor_str,
        lst.price or 0,
        lst.deposit,
        lst.contract_months,
        lst.amenities,
        lst.description[:200] if lst.description else "",
        len(lst.photos),
        lst.landlord_phone,
        lst.landlord_username,
        lst.created_at,
        lst.approved_at,
        lst.posted_count,
        lst.rented_at,
    ]


# ── Main export function ──────────────────────────────────────────────────────

def export_listings_to_xlsx(
    listings: List[Listing],
    filter_status: Optional[ListingStatus] = None,
    title: str = "Astana Rental Listings",
) -> bytes:
    """
    Exports a list of Listing objects to a styled .xlsx workbook.

    Returns raw bytes ready to be sent via Bot.send_document() without
    writing any temporary files to disk.

    Args:
        listings:      All listings to include (filtered if filter_status is set).
        filter_status: If provided, only listings with this status are exported.
        title:         Workbook title displayed in cell A1.
    """
    if filter_status:
        listings = [lst for lst in listings if lst.status == filter_status]

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    # ── Document title (row 1) ─────────────────────────────────────────────────
    last_col = get_column_letter(len(_COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Export metadata (row 2) ────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Exported: {datetime.now().strftime('%d.%m.%Y %H:%M')} | "
        f"Records: {len(listings)}"
    )
    ws["A2"].font = Font(size=9, name="Arial", color="888888")
    ws.row_dimensions[2].height = 16

    # Row 3 is intentionally empty (spacer)
    ws.row_dimensions[3].height = 8

    # ── Column headers (row 4) ─────────────────────────────────────────────────
    for col_idx, (col_name, col_width, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[4].height = 24

    # ── Data rows (starting at row 5) ─────────────────────────────────────────
    for row_offset, lst in enumerate(listings, start=5):
        values  = _listing_to_row(lst)
        status_fill = _STATUS_FILL.get(lst.status)
        is_even     = row_offset % 2 == 0

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font      = _BODY_FONT
            cell.border    = _BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx in _WRAP_COL_INDICES),
            )

            # Background: status colour takes priority, then alternating rows
            if status_fill:
                cell.fill = status_fill
            elif is_even:
                cell.fill = _ALT_ROW_FILL

            if col_idx in _MONEY_COL_INDICES and isinstance(value, (int, float)):
                cell.number_format = _MONEY_FORMAT

            if col_idx in _DATE_COL_INDICES and isinstance(value, datetime):
                cell.number_format = _DATE_FORMAT

        ws.row_dimensions[row_offset].height = 20

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A5"

    # ── Additional worksheets ──────────────────────────────────────────────────
    _write_stats_sheet(wb.create_sheet("Statistics"), listings)
    _write_legend_sheet(wb.create_sheet("Legend"))

    # Serialise to bytes (no temp file)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Statistics worksheet ──────────────────────────────────────────────────────

def _write_stats_sheet(ws, listings: List[Listing]) -> None:
    """Writes a summary statistics tab to the given worksheet."""
    ws["A1"] = "Listing statistics"
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1A3A5C")

    total     = len(listings)
    by_status: dict = {}
    by_rooms:  dict = {}

    for lst in listings:
        by_status[lst.status.value] = by_status.get(lst.status.value, 0) + 1
        if lst.rooms == 0:
            key = "Studio"
        elif lst.rooms:
            key = f"{lst.rooms}-room"
        else:
            key = "Not specified"
        by_rooms[key] = by_rooms.get(key, 0) + 1

    bold = Font(bold=True, name="Arial")
    row  = 3

    ws.cell(row=row, column=1, value="Total listings").font = bold
    ws.cell(row=row, column=2, value=total)
    row += 1

    last_data_row = 4 + total
    ws.cell(row=row, column=1, value="Average price ₸/month").font = bold
    ws.cell(row=row, column=2, value=f"=AVERAGE(Listings!H5:H{last_data_row})")
    ws.cell(row=row, column=2).number_format = _MONEY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Min price").font = bold
    ws.cell(row=row, column=2, value=f"=MIN(Listings!H5:H{last_data_row})")
    ws.cell(row=row, column=2).number_format = _MONEY_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Max price").font = bold
    ws.cell(row=row, column=2, value=f"=MAX(Listings!H5:H{last_data_row})")
    ws.cell(row=row, column=2).number_format = _MONEY_FORMAT
    row += 2

    ws.cell(row=row, column=1, value="By status:").font = bold
    row += 1
    for status_val, count in by_status.items():
        ws.cell(row=row, column=1, value=status_val)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="By rooms:").font = bold
    row += 1
    for rooms_key, count in sorted(by_rooms.items()):
        ws.cell(row=row, column=1, value=rooms_key)
        ws.cell(row=row, column=2, value=count)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18


# ── Legend worksheet ──────────────────────────────────────────────────────────

def _write_legend_sheet(ws) -> None:
    """Writes a colour legend tab to the given worksheet."""
    ws["A1"] = "Row colour legend"
    ws["A1"].font = Font(bold=True, size=12, name="Arial")

    legend = [
        ("Pending",   "FFF3CD", "Awaiting admin approval"),
        ("Published", "D4EDDA", "Active listing in the channel"),
        ("Rented",    "D6D8DB", "Apartment rented — listing inactive"),
        ("Rejected",  "F8D7DA", "Rejected by moderator"),
    ]
    for i, (label, color, desc) in enumerate(legend, start=3):
        ws.cell(row=i, column=1, value=label).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=i, column=2, value=desc)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40


# ── Telegram delivery helper ──────────────────────────────────────────────────

async def send_xlsx_to_admin(bot, admin_id: int, listings: List[Listing]) -> None:
    """Generates the report and sends it as a document to the given admin."""

    xlsx_bytes = export_listings_to_xlsx(listings)
    filename   = f"astana_rent_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    await bot.send_document(
        chat_id=admin_id,
        document=BufferedInputFile(xlsx_bytes, filename=filename),
        caption=(
            f"📊 <b>Listings export</b>\n"
            f"Total: {len(listings)}\n"
            f"Date: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ),
        parse_mode="HTML",
    )
